"""
export_report.py — ייצוא דוח Excel מקצועי
3 גיליונות: ניתוח משחק + Value Bets + Elo Rankings
"""

import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ─── צבעים ─────────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1E3A6E"   # כחול כהה
C_HEADER_FG  = "FFFFFF"   # לבן
C_VALUE_BG   = "0D4A2E"   # ירוק כהה — Value Bet
C_VALUE_FG   = "10B981"   # ירוק בהיר
C_NO_VALUE   = "3B1C1C"   # אדום כהה
C_SUBHEAD    = "111827"   # שורת כותרת משנה
C_ROW_ALT    = "0F1927"   # שורות חלופיות
C_ACCENT     = "3B82F6"   # כחול בהיר
C_GOLD       = "F59E0B"   # זהב — מדד Elo גבוה


def _header_style(cell, bg=C_HEADER_BG, fg=C_HEADER_FG, size=11, bold=True):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _border():
    s = Side(style="thin", color="1F2D4A")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def build_excel_report(
    match_data: dict | None,
    value_bets: list[dict],
    elo_rankings: list[dict],
) -> bytes:
    """
    בונה דוח Excel ומחזיר bytes להורדה.

    match_data: {
        home_name, away_name, match_date, venue, city,
        elo_home, elo_away, form_home, form_away,
        xg_home, xg_away,
        probs: {home, draw, away},
        fair_odds: {home, draw, away},
        live_odds: {home, draw, away, bookmaker} | None,
        ev: {home, draw, away},
        kelly: {home, draw, away},
        top_scores: [(score, pct), ...],
        injuries_home: [str], injuries_away: [str],
        h2h: [{date, home, result, away}, ...]
    }
    """
    wb = Workbook()

    # ══════════════════════════════════════════════════════
    # גיליון 1 — Match Analysis
    # ══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Match Analysis"
    ws1.sheet_view.rightToLeft = True

    if match_data:
        row = 1
        # כותרת ראשית
        ws1.merge_cells(f"A{row}:H{row}")
        c = ws1.cell(row, 1, f"🏆 {match_data['home_name']} נגד {match_data['away_name']}")
        _header_style(c, size=14)
        ws1.row_dimensions[row].height = 30
        row += 1

        ws1.merge_cells(f"A{row}:H{row}")
        c = ws1.cell(row, 1, f"{match_data['match_date']} · {match_data.get('venue','')} · {match_data.get('city','')}")
        _header_style(c, bg="0D1B3E", size=10, bold=False)
        ws1.row_dimensions[row].height = 18
        row += 2

        # ── נתוני מפתח ──
        ws1.merge_cells(f"A{row}:H{row}")
        c = ws1.cell(row, 1, "📊 נתוני מפתח")
        _header_style(c, bg=C_SUBHEAD, size=11)
        row += 1

        key_data = [
            ["מדד Elo", match_data['home_name'], match_data['away_name']],
            ["", match_data['elo_home'], match_data['elo_away']],
            ["טופס (5 משחקים)", f"{match_data['form_home']:.2f}x", f"{match_data['form_away']:.2f}x"],
            ["xG צפוי", match_data['xg_home'], match_data['xg_away']],
        ]
        for krow in key_data:
            for col, val in enumerate(krow, 1):
                c = ws1.cell(row, col, val)
                c.font = Font(name="Arial", size=10, color="C8D0E0")
                c.alignment = Alignment(horizontal="center")
                c.border = _border()
            row += 1
        row += 1

        # ── הסתברויות ועמדת ערך ──
        ws1.merge_cells(f"A{row}:H{row}")
        c = ws1.cell(row, 1, "🎯 תחזית המודל")
        _header_style(c, bg=C_SUBHEAD, size=11)
        row += 1

        headers = ["תוצאה", "סיכוי %", "יחס הוגן", "Odds (אתר)", "EV", "Kelly %", "Value?"]
        for col, h in enumerate(headers, 1):
            c = ws1.cell(row, col, h)
            _header_style(c, size=10)
        row += 1

        outcomes = [
            (f"{match_data['home_name']} מנצחת", "home"),
            ("תיקו", "draw"),
            (f"{match_data['away_name']} מנצחת", "away"),
        ]
        for label, key in outcomes:
            prob   = match_data['probs'].get(key, 0)
            fair   = match_data['fair_odds'].get(key, 0)
            live   = match_data.get('live_odds', {})
            odd    = live.get(key, "-") if live else "-"
            ev     = match_data['ev'].get(key, 0)
            kelly  = match_data['kelly'].get(key, 0)
            is_val = ev > 0

            row_vals = [label, f"{prob:.1f}%", fair, odd,
                        f"+{ev:.1%}" if ev > 0 else f"{ev:.1%}",
                        f"{kelly:.1f}%" if kelly > 0 else "-",
                        "✅ VALUE" if is_val else "❌"]

            bg = C_VALUE_BG if is_val else C_ROW_ALT
            fg = C_VALUE_FG if is_val else "C8D0E0"

            for col, val in enumerate(row_vals, 1):
                c = ws1.cell(row, col, val)
                c.font = Font(name="Arial", size=10, color=fg, bold=is_val)
                c.fill = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(horizontal="center")
                c.border = _border()
            row += 1
        row += 1

        # ── תוצאות סבירות ──
        ws1.merge_cells(f"A{row}:H{row}")
        c = ws1.cell(row, 1, "⚽ תוצאות סבירות ביותר")
        _header_style(c, bg=C_SUBHEAD, size=11)
        row += 1

        for col, (score, pct) in enumerate(match_data.get('top_scores', [])[:5], 1):
            ws1.cell(row, col, score).font = Font(name="Arial", bold=True, color=C_ACCENT, size=12)
            ws1.cell(row, col).alignment = Alignment(horizontal="center")
            ws1.cell(row+1, col, f"{pct}%").font = Font(name="Arial", size=9, color="6B7A99")
            ws1.cell(row+1, col).alignment = Alignment(horizontal="center")
        row += 3

        # ── פציעות ──
        if match_data.get('injuries_home') or match_data.get('injuries_away'):
            ws1.merge_cells(f"A{row}:H{row}")
            c = ws1.cell(row, 1, "🚑 פצועים ונעדרים")
            _header_style(c, bg=C_SUBHEAD, size=11)
            row += 1
            for name in (match_data.get('injuries_home') or ["ללא נפגעים"]):
                ws1.cell(row, 1, f"🤕 {name} ({match_data['home_name']})").font = Font(name="Arial", size=10, color="FCA5A5")
                row += 1
            for name in (match_data.get('injuries_away') or ["ללא נפגעים"]):
                ws1.cell(row, 1, f"🤕 {name} ({match_data['away_name']})").font = Font(name="Arial", size=10, color="FCA5A5")
                row += 1
            row += 1

        # ── H2H ──
        if match_data.get('h2h'):
            ws1.merge_cells(f"A{row}:H{row}")
            c = ws1.cell(row, 1, "⚔️ עימותים ישירים")
            _header_style(c, bg=C_SUBHEAD, size=11)
            row += 1
            for g in match_data['h2h']:
                ws1.cell(row, 1, g.get('date','')).font = Font(name="Arial", size=9, color="6B7A99")
                ws1.cell(row, 2, g.get('home','')).font = Font(name="Arial", size=10, color="C8D0E0")
                ws1.cell(row, 3, g.get('result','')).font = Font(name="Arial", size=10, bold=True, color=C_ACCENT)
                ws1.cell(row, 3).alignment = Alignment(horizontal="center")
                ws1.cell(row, 4, g.get('away','')).font = Font(name="Arial", size=10, color="C8D0E0")
                row += 1

        # רוחב עמודות
        for col, w in zip("ABCDEFGH", [28, 12, 12, 14, 10, 10, 10, 10]):
            _set_col_width(ws1, col, w)

        ws1.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════════
    # גיליון 2 — Value Bets
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Value Bets")
    ws2.sheet_view.rightToLeft = True

    ws2.merge_cells("A1:J1")
    c = ws2.cell(1, 1, f"💰 Value Bets — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    _header_style(c, size=13)
    ws2.row_dimensions[1].height = 28

    vb_headers = ["תאריך", "משחק", "הימור על", "Odds", "סיכוי %", "סיכוי משתמע %", "EV", "Kelly %", "Overround %", "אתר"]
    for col, h in enumerate(vb_headers, 1):
        c = ws2.cell(2, col, h)
        _header_style(c, size=10)

    for i, vb in enumerate(value_bets, 3):
        ev = vb.get("EV", 0)
        is_high = ev > 0.10
        bg = "0A2A1A" if is_high else C_ROW_ALT
        fg = "6EE7B7" if is_high else "C8D0E0"

        row_vals = [
            vb.get("תאריך", ""),
            vb.get("משחק", ""),
            vb.get("הימור על", ""),
            vb.get("Odds", ""),
            f"{vb.get('סיכוי %', 0):.1f}%",
            f"{vb.get('סיכוי משתמע %', 0):.1f}%",
            f"+{ev:.1%}" if ev > 0 else f"{ev:.1%}",
            f"{vb.get('Kelly %', 0):.1f}%",
            f"{vb.get('Overround %', 0):.1f}%",
            vb.get("אתר", ""),
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws2.cell(i, col, val)
            c.font = Font(name="Arial", size=10, color=fg)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center")
            c.border = _border()

    for col, w in zip("ABCDEFGHIJ", [12, 30, 20, 8, 10, 14, 8, 8, 10, 12]):
        _set_col_width(ws2, col, w)
    ws2.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════════
    # גיליון 3 — Elo Rankings
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Elo Rankings")
    ws3.sheet_view.rightToLeft = True

    ws3.merge_cells("A1:C1")
    c = ws3.cell(1, 1, "📊 דירוג Elo — מונדיאל 2026")
    _header_style(c, size=13)
    ws3.row_dimensions[1].height = 28

    for col, h in enumerate(["#", "נבחרת", "מדד Elo"], 1):
        c = ws3.cell(2, col, h)
        _header_style(c, size=10)

    for i, team in enumerate(elo_rankings, 1):
        elo = team.get("elo_rating", 0)
        if elo >= 1700:
            fg, bg = C_GOLD, "1A1200"
        elif elo >= 1600:
            fg, bg = C_ACCENT, "0A1520"
        else:
            fg, bg = "C8D0E0", C_ROW_ALT

        ws3.cell(i+2, 1, i).font = Font(name="Arial", size=10, color="6B7A99")
        ws3.cell(i+2, 1).alignment = Alignment(horizontal="center")

        c = ws3.cell(i+2, 2, team.get("name", ""))
        c.font = Font(name="Arial", size=10, color=fg, bold=(elo >= 1700))
        c.fill = PatternFill("solid", start_color=bg)

        c = ws3.cell(i+2, 3, elo)
        c.font = Font(name="Arial", size=11, color=fg, bold=(elo >= 1700))
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center")
        c.border = _border()

    for col, w in zip("ABC", [6, 24, 12]):
        _set_col_width(ws3, col, w)
    ws3.sheet_view.showGridLines = False

    # ── שמירה ל-bytes ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()